import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\Usuario\Downloads\PB_Vasquez_Augusto_Reyes_Renato (1).xlsx"
DST = r"C:\Users\Usuario\Downloads\PB_Sinapsistencia_v2.xlsx"

wb = load_workbook(SRC)

# ── helpers ──────────────────────────────────────────────────────────────────
def cell(ws, r, c, value=None, bold=False, size=11, color="000000",
         bg=None, align="left", valign="center", wrap=False, italic=False):
    cel = ws.cell(row=r, column=c)
    if value is not None:
        cel.value = value
    cel.font = Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
    cel.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if bg:
        cel.fill = PatternFill("solid", fgColor=bg)
    return cel

def thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

# ── Paleta ───────────────────────────────────────────────────────────────────
NAVY   = "0D1B2A"
TEAL   = "1B6CA8"
LTGRAY = "F5F5F5"
DGRAY  = "4A4A4A"
WHITE  = "FFFFFF"
MUST   = "C0392B"
SHOULD = "E67E22"
COULD  = "D4AC0D"
DONE_BG = "D5F5E3"; DONE_FG = "1E8449"
PEND_BG = "FEF9E7"; PEND_FG = "9A7D0A"
PLAN_BG = "EBF5FB"; PLAN_FG = "1A5276"

# ════════════════════════════════════════════════════════════════════
# SHEET 1 — PRODUCT BACKLOG
# ════════════════════════════════════════════════════════════════════
old = wb.sheetnames[0]
del wb[old]
ws = wb.create_sheet("Product Backlog", 0)

widths = {1:6, 2:10, 3:44, 4:13, 5:5, 6:12, 7:9, 8:15, 9:28}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# Cabecera
merge(ws, 1, 1, 1, 9)
cell(ws, 1, 1, "SINAPSISTENCIA — PRODUCT BACKLOG",
     bold=True, size=16, color=WHITE, bg=NAVY, align="center")
ws.row_dimensions[1].height = 28

merge(ws, 2, 1, 2, 9)
meta = ("Plataforma web ML para asesoramiento medico-legal | Caso: Clinica SANNA \"El Golf\" | "
        "PO: Vasquez Requejo A. & Reyes Valenzuela R. | Equipo: 2 devs | 7 Sprints x 2 semanas")
cell(ws, 2, 1, meta, size=9, color=WHITE, bg=TEAL, align="center", wrap=True)
ws.row_dimensions[2].height = 28

merge(ws, 3, 1, 3, 9)
cell(ws, 3, 1,
     "LEYENDA:  Rojo=Must Have  Naranja=Should Have  Amarillo=Could Have  |  Verde=Completado  Amarillo=Pendiente  Azul=Planificado",
     size=9, color=DGRAY, bg="F0F0F0", align="center", italic=True)
ws.row_dimensions[3].height = 16

# Cabecera tabla
HDR = 4
headers = ["#", "ID", "Historia de Usuario", "Epica", "SP", "Prioridad", "Sprint", "Estado", "Dependencias"]
for c, h in enumerate(headers, 1):
    cel = cell(ws, HDR, c, h, bold=True, size=10, color=WHITE, bg=NAVY, align="center")
    cel.border = thin_border()
ws.row_dimensions[HDR].height = 24

# Datos backlog
backlog = [
    # (num, id, historia, epica, sp, prio, sprint, estado, deps)
    ("E1", None, "EPICA 1: Autenticacion y Control de Acceso", None, None, None, None, None, None),
    (1,"HU-001","Inicio de sesion por credenciales","Auth",5,"Must",1,"Completado","—"),
    (2,"HU-002","Cierre de sesion","Auth",3,"Must",1,"Completado","HU-001"),
    (3,"HU-003","Proteccion de rutas por rol","Auth",8,"Must",1,"Completado","HU-001"),
    (4,"HU-004","Acceso rapido por rol (modo demo)","Auth",3,"Should",1,"Completado","HU-001"),
    (5,"HU-005","Cambio de rol desde sidebar (testing)","Auth",3,"Could",1,"Completado","HU-001, HU-003"),
    ("E10",None,"EPICA 10: Landing y Registro",None,None,None,None,None,None),
    (6,"HU-032","Pagina de inicio (landing)","Landing",5,"Must",1,"Completado","—"),
    (7,"HU-033","Formulario de solicitud de acceso","Landing",5,"Should",1,"Completado","—"),
    ("E2",None,"EPICA 2: Gestion Documental Clinico-Legal",None,None,None,None,None,None),
    (8,"HU-006","Listado de documentos con filtros","Documents",5,"Must",2,"Completado","HU-001, HU-003"),
    (9,"HU-007","Creacion de documentos","Documents",5,"Must",2,"Completado","HU-006"),
    (10,"HU-008","Detalle y trazabilidad de un documento","Documents",5,"Must",2,"Completado","HU-006"),
    (11,"HU-009","Cambio de estado de un documento","Documents",3,"Must",2,"Completado","HU-008"),
    (12,"HU-010","Listado de documentos (vista admin)","Documents",3,"Should",4,"Completado","HU-006"),
    (13,"HU-041","Historial de versiones de un documento","Documents",5,"Should",5,"Pendiente","HU-008"),
    (14,"HU-042","Busqueda y filtrado avanzado de documentos","Documents",5,"Could",5,"Pendiente","HU-006"),
    ("E3",None,"EPICA 3: Gestion de Casos Legales (exclusivo medico-abogado)",None,None,None,None,None,None),
    (15,"HU-011","Listado de casos del medico","Cases",5,"Must",3,"Completado","HU-001, HU-003"),
    (16,"HU-012","Creacion de un caso legal","Cases",5,"Must",3,"Completado","HU-011"),
    (17,"HU-013","Detalle de un caso legal","Cases",5,"Must",3,"Completado","HU-011"),
    (18,"HU-014","Gestion de episodios clinicos (admin)","Cases",8,"Should",4,"Completado","HU-003"),
    (19,"HU-037","Asignacion de abogado a un caso legal","Cases",5,"Should",5,"Pendiente","HU-013"),
    (20,"HU-038","Edicion de notas de un caso legal","Cases",3,"Could",5,"Pendiente","HU-013"),
    ("E4",None,"EPICA 4: Vinculacion Medico-Abogado (Matching)",None,None,None,None,None,None),
    (21,"HU-015","Recomendaciones de abogados para medico","Matching",5,"Must",3,"Completado","HU-001"),
    (22,"HU-016","Envio de solicitud de contacto","Matching",5,"Must",3,"Completado","HU-015"),
    (23,"HU-017","Gestion de solicitudes recibidas (abogado)","Matching",5,"Must",3,"Completado","HU-016"),
    (24,"HU-019","Directorio de abogados (vista medico)","Matching",5,"Should",3,"Completado","HU-001"),
    (25,"HU-020","Directorio de medicos (vista abogado)","Matching",5,"Should",3,"Completado","HU-001"),
    (26,"HU-039","Mensaje de respuesta al gestionar solicitudes","Matching",3,"Could",6,"Pendiente","HU-017"),
    ("E5",None,"EPICA 5: Gestion de Pacientes",None,None,None,None,None,None),
    (27,"HU-021","Listado de pacientes","Patients",5,"Must",4,"Completado","HU-003"),
    (28,"HU-022","Registro de nuevo paciente","Patients",5,"Must",4,"Completado","HU-021"),
    (29,"HU-023","Edicion de paciente","Patients",3,"Should",4,"Completado","HU-021"),
    ("E6",None,"EPICA 6: Gestion de Usuarios",None,None,None,None,None,None),
    (30,"HU-024","Listado de usuarios del sistema","Users",5,"Must",4,"Completado","HU-003"),
    (31,"HU-025","Creacion de nuevo usuario","Users",5,"Must",4,"Completado","HU-024"),
    (32,"HU-026","Activar/desactivar usuario","Users",2,"Should",4,"Completado","HU-024"),
    ("E7",None,"EPICA 7: Auditoria y Trazabilidad",None,None,None,None,None,None),
    (33,"HU-027","Registro de auditoria","Audit",5,"Must",4,"Completado","HU-003"),
    (34,"HU-043","Exportar listado de auditoria a CSV","Audit",3,"Could",6,"Pendiente","HU-027"),
    ("E8",None,"EPICA 8: Dashboards y Visualizacion",None,None,None,None,None,None),
    (35,"HU-028","Dashboard del medico","Dashboard",8,"Must",2,"Completado","HU-001, HU-003"),
    (36,"HU-018","Dashboard del abogado con solicitudes","Dashboard",8,"Must",3,"Completado","HU-017"),
    (37,"HU-029","Dashboard del administrador","Dashboard",8,"Must",4,"Completado","HU-024, HU-027"),
    ("E9",None,"EPICA 9: Perfil de Usuario",None,None,None,None,None,None),
    (38,"HU-030","Visualizacion y edicion de perfil del medico","Profile",5,"Should",5,"Completado","HU-001"),
    (39,"HU-031","Visualizacion y edicion de perfil del abogado","Profile",5,"Should",5,"Completado","HU-001"),
    (40,"HU-040","Toggle de disponibilidad del abogado","Profile",2,"Could",6,"Pendiente","HU-031"),
    ("E11",None,"EPICA 11: Firmas Digitales y Consentimiento Informado",None,None,None,None,None,None),
    (41,"HU-034","Firma de un documento por parte del medico","Signatures",5,"Should",5,"Pendiente","HU-008, HU-009"),
    (42,"HU-035","Gestion de registros de consentimiento informado","Signatures",8,"Should",6,"Pendiente","HU-034"),
    ("E12",None,"EPICA 12: Gestion de Archivos Adjuntos",None,None,None,None,None,None),
    (43,"HU-036","Adjuntar documentos a un caso legal","Attachments",8,"Should",6,"Pendiente","HU-013"),
    ("E13",None,"EPICA 13: Evaluacion de Riesgo ML — Sprint 7 (Planificado)",None,None,None,None,None,None),
    (44,"HU-044","Evaluacion de riesgo medico-legal de un caso","ML-Risk",13,"Must",7,"Planificado","HU-013"),
    (45,"HU-045","Visualizacion de factores de riesgo (XAI)","ML-Risk",8,"Must",7,"Planificado","HU-044"),
    (46,"HU-046","Historial de evaluaciones de riesgo por paciente","ML-Risk",5,"Should",7,"Planificado","HU-044"),
    (47,"HU-047","Recomendaciones automaticas basadas en riesgo","ML-Risk",5,"Should",7,"Planificado","HU-044"),
]

ROW = HDR + 1
for entry in backlog:
    num = entry[0]
    is_epic = isinstance(num, str) and num.startswith("E")

    if is_epic:
        ws.row_dimensions[ROW].height = 20
        merge(ws, ROW, 1, ROW, 9)
        cell(ws, ROW, 1, entry[2], bold=True, size=10, color=WHITE, bg=TEAL, align="left")
        ROW += 1
        continue

    _, hid, historia, epica, sp, prio, spr, estado, deps = entry
    ws.row_dimensions[ROW].height = 18
    bg = LTGRAY if ROW % 2 == 0 else WHITE

    cell(ws, ROW, 1, num, align="center", size=9, color=DGRAY, bg=bg).border = thin_border()
    cell(ws, ROW, 2, hid, bold=True, size=9, color=TEAL, align="center", bg=bg).border = thin_border()
    cell(ws, ROW, 3, historia, size=10, wrap=True, bg=bg).border = thin_border()
    cell(ws, ROW, 4, epica, size=9, color=DGRAY, align="center", bg=bg).border = thin_border()
    cell(ws, ROW, 5, sp, bold=True, size=11, color=NAVY, align="center", bg=bg).border = thin_border()

    prio_map = {"Must":(MUST,WHITE,"Must Have"), "Should":(SHOULD,WHITE,"Should Have"), "Could":(COULD,WHITE,"Could Have")}
    pbg, pfg, plbl = prio_map.get(prio, ("AAAAAA",WHITE,prio))
    cell(ws, ROW, 6, plbl, bold=True, size=8, color=pfg, bg=pbg, align="center").border = thin_border()

    cell(ws, ROW, 7, f"S{spr}", size=9, color=DGRAY, align="center", bg=bg).border = thin_border()

    if estado == "Completado":
        ebg, efg, elbl = DONE_BG, DONE_FG, "Completado"
    elif estado == "Planificado":
        ebg, efg, elbl = PLAN_BG, PLAN_FG, "Planificado"
    else:
        ebg, efg, elbl = PEND_BG, PEND_FG, "Pendiente"
    cell(ws, ROW, 8, elbl, bold=True, size=9, color=efg, bg=ebg, align="center").border = thin_border()

    cell(ws, ROW, 9, deps, size=9, color=DGRAY, wrap=True, bg=bg).border = thin_border()
    ROW += 1

# Totales
ROW += 1
merge(ws, ROW, 1, ROW, 4)
cell(ws, ROW, 1, "TOTAL SP (sin ML):", bold=True, size=10, color=NAVY, align="right")
cell(ws, ROW, 5, 212, bold=True, size=12, color=WHITE, bg=NAVY, align="center")
merge(ws, ROW, 6, ROW, 7)
cell(ws, ROW, 6, "TOTAL SP (con ML):", bold=True, size=10, color=NAVY, align="right")
cell(ws, ROW, 8, 243, bold=True, size=12, color=WHITE, bg=TEAL, align="center")
cell(ws, ROW, 9, "47 HUs | 13 Epicas | 7 Sprints | 2 devs", size=9, color=DGRAY, align="center")
ws.row_dimensions[ROW].height = 22

# Resumen sprints
ROW += 3
merge(ws, ROW, 1, ROW, 9)
cell(ws, ROW, 1, "RESUMEN POR SPRINT", bold=True, size=12, color=WHITE, bg=NAVY, align="center")
ws.row_dimensions[ROW].height = 22
ROW += 1
for c, h in zip([1,2,3,4,5,6,7,8], ["Sprint","Semanas","Objetivo","HUs","SP Planif.","SP Complet.","Velocidad","Estado"]):
    cell(ws, ROW, c, h, bold=True, size=9, color=WHITE, bg=TEAL, align="center").border = thin_border()
ws.row_dimensions[ROW].height = 18

sprints = [
    ("S1","13-24 Ene","Auth + Landing + Registro",7,32,32,"32 SP","Completado"),
    ("S2","27 Ene-7 Feb","Gestion documental + Dashboard medico",5,26,26,"26 SP","Completado"),
    ("S3","10-21 Feb","Casos legales + Matching + Dashboards",9,48,48,"48 SP","Completado"),
    ("S4","24 Feb-7 Mar","Panel administrativo completo",10,49,49,"49 SP","Completado"),
    ("S5","10-21 Mar","Perfiles + Firmas + Mejoras",7,33,0,"—","Pendiente"),
    ("S6","24 Mar-4 Abr","Consentimiento + Adjuntos + CSV",5,24,0,"—","Pendiente"),
    ("S7","7-18 Abr","Modulo de Machine Learning",4,31,0,"—","Planificado"),
]
ROW += 1
for s in sprints:
    ws.row_dimensions[ROW].height = 18
    bg = LTGRAY if ROW % 2 == 0 else WHITE
    sn,sw,so,sh,ssp,ssc,sv,sst = s
    cell(ws,ROW,1,sn,bold=True,size=9,color=TEAL,align="center",bg=bg).border=thin_border()
    cell(ws,ROW,2,sw,size=9,color=DGRAY,align="center",bg=bg).border=thin_border()
    cell(ws,ROW,3,so,size=9,wrap=True,bg=bg).border=thin_border()
    cell(ws,ROW,4,sh,bold=True,size=10,color=NAVY,align="center",bg=bg).border=thin_border()
    cell(ws,ROW,5,ssp,bold=True,size=10,color=NAVY,align="center",bg=bg).border=thin_border()
    cell(ws,ROW,6,ssc if ssc else "—",bold=True,size=10,color=NAVY,align="center",bg=bg).border=thin_border()
    cell(ws,ROW,7,sv,size=9,color=DGRAY,align="center",bg=bg).border=thin_border()
    if sst=="Completado": ebg,efg,el=DONE_BG,DONE_FG,"Completado"
    elif sst=="Planificado": ebg,efg,el=PLAN_BG,PLAN_FG,"Planificado"
    else: ebg,efg,el=PEND_BG,PEND_FG,"Pendiente"
    cell(ws,ROW,8,el,bold=True,size=9,color=efg,bg=ebg,align="center").border=thin_border()
    ROW += 1

# Riesgos
ROW += 2
merge(ws,ROW,1,ROW,9)
cell(ws,ROW,1,"RIESGOS IDENTIFICADOS",bold=True,size=12,color=WHITE,bg=NAVY,align="center")
ws.row_dimensions[ROW].height=22
ROW+=1
for c,h in zip([1,2,3,4,5,6,7,8,9],["ID","Descripcion del riesgo","","","","Prob.","Impacto","Mitigacion",""]):
    cell(ws,ROW,c,h,bold=True,size=9,color=WHITE,bg=TEAL,align="center").border=thin_border()
ws.row_dimensions[ROW].height=18
risks=[
    ("R1","Retraso ML por complejidad del modelo","Alta","Alto","Implementar con datos simulados; backend real en iteracion posterior"),
    ("R2","Cambios en requisitos del caso de estudio","Media","Medio","Arquitectura modular permite adaptar sin reescribir modulos"),
    ("R3","Dependencias tecnicas entre modulos causan bloqueos","Baja","Medio","Sprint planning con analisis de dependencias; DI container"),
    ("R4","Modelo ML con precision < 85%","Media","Alto","Iterar multiples algoritmos; usar ensemble methods"),
    ("R5","Incompatibilidad con regulaciones MINSA/SUSALUD","Baja","Alto","Validar con marco legal Cap.2; modulo auditoria cumple trazabilidad"),
]
prob_c={"Alta":"C0392B","Media":"E67E22","Baja":"27AE60"}
ROW+=1
for r in risks:
    ws.row_dimensions[ROW].height=18
    rid,rdesc,rprob,rimp,rmit=r
    bg=LTGRAY if ROW%2==0 else WHITE
    cell(ws,ROW,1,rid,bold=True,size=9,color=TEAL,align="center",bg=bg).border=thin_border()
    merge(ws,ROW,2,ROW,5)
    cell(ws,ROW,2,rdesc,size=9,wrap=True,bg=bg).border=thin_border()
    cell(ws,ROW,6,rprob,bold=True,size=9,color=WHITE,bg=prob_c.get(rprob,"888888"),align="center").border=thin_border()
    cell(ws,ROW,7,rimp,bold=True,size=9,color=WHITE,bg=prob_c.get(rimp,"888888"),align="center").border=thin_border()
    merge(ws,ROW,8,ROW,9)
    cell(ws,ROW,8,rmit,size=9,wrap=True,bg=bg).border=thin_border()
    ROW+=1

ws.freeze_panes="A5"
ws.auto_filter.ref=f"A{HDR}:I{HDR}"

# ════════════════════════════════════════════════════════════════════
# SHEET 2 — HISTORIAS DE USUARIO (nuevas HUs del proyecto)
# ════════════════════════════════════════════════════════════════════
old2 = wb.sheetnames[1]
del wb[old2]
ws2 = wb.create_sheet("Historias de Usuario", 1)

# Anchos hoja 2
ws2_widths = {1:10, 2:45, 3:38, 4:22, 5:12, 6:8}
for col, w in ws2_widths.items():
    ws2.column_dimensions[get_column_letter(col)].width = w

# Cabecera hoja 2
merge(ws2, 1, 1, 1, 6)
cell(ws2, 1, 1, "HISTORIAS DE USUARIO — SINAPSISTENCIA",
     bold=True, size=14, color=WHITE, bg=NAVY, align="center")
ws2.row_dimensions[1].height = 26

merge(ws2, 2, 1, 2, 6)
cell(ws2, 2, 1,
     "Clinica SANNA \"El Golf\" | 43 historias de usuario + 4 planificadas ML | 13 Epicas",
     size=9, color=WHITE, bg=TEAL, align="center", italic=True)
ws2.row_dimensions[2].height = 18

# Cabecera tabla HU
HDR2 = 3
hu_headers = ["ID", "Nombre de la Historia de Usuario", "Quiero...", "Rol", "Prioridad", "Sprint"]
for c, h in enumerate(hu_headers, 1):
    cel = cell(ws2, HDR2, c, h, bold=True, size=10, color=WHITE, bg=NAVY, align="center")
    cel.border = thin_border()
ws2.row_dimensions[HDR2].height = 22

# Datos HUs (todas las 47 HUs del proyecto actual)
hu_data = [
    # Epica 1
    ("epica", "EPICA 1: Autenticacion y Control de Acceso"),
    ("HU-001","Inicio de sesion por credenciales","iniciar sesion con mi correo y contrasena para acceder de forma segura segun mi rol","Medico / Abogado / Admin","Must Have",1),
    ("HU-002","Cierre de sesion","cerrar mi sesion de forma segura eliminando la cookie y el estado","Medico / Abogado / Admin","Must Have",1),
    ("HU-003","Proteccion de rutas por rol","que cada rol solo pueda acceder a las rutas de su portal","Admin del sistema","Must Have",1),
    ("HU-004","Acceso rapido por rol (modo demo)","poder iniciar sesion rapidamente seleccionando un rol sin credenciales","Evaluador / Revisor","Should Have",1),
    ("HU-005","Cambio de rol desde sidebar (testing)","cambiar de rol sin cerrar sesion desde el sidebar","Evaluador del sistema","Could Have",1),
    # Epica 10
    ("epica", "EPICA 10: Landing y Registro"),
    ("HU-032","Pagina de inicio (landing)","ver una pagina de inicio que explique que es Sinapsistencia y sus funcionalidades","Visitante no autenticado","Must Have",1),
    ("HU-033","Formulario de solicitud de acceso","solicitar acceso a la plataforma proporcionando mis datos profesionales","Profesional de salud / Abogado","Should Have",1),
    # Epica 2
    ("epica", "EPICA 2: Gestion Documental Clinico-Legal"),
    ("HU-006","Listado de documentos con filtros","ver todos mis documentos clinico-legales en una tabla con busqueda y filtros","Medico","Must Have",2),
    ("HU-007","Creacion de documentos","crear un nuevo documento clinico-legal seleccionando tipo, paciente y contenido","Medico","Must Have",2),
    ("HU-008","Detalle y trazabilidad de un documento","ver el detalle completo de un documento incluyendo versiones, firmas y metadatos","Medico / Admin","Must Have",2),
    ("HU-009","Cambio de estado de un documento","cambiar el estado de un documento (borrador, pendiente firma, firmado, archivado)","Medico","Must Have",2),
    ("HU-010","Listado de documentos (vista admin)","ver todos los documentos del sistema para supervisar la documentacion institucional","Administrador","Should Have",4),
    ("HU-041","Historial de versiones de un documento","ver el historial completo de versiones y consultar el contenido de cada version","Medico / Admin","Should Have",5),
    ("HU-042","Busqueda y filtrado avanzado de documentos","filtrar documentos por tipo, estado y rango de fechas ademas de busqueda por texto","Medico / Admin","Could Have",5),
    # Epica 3
    ("epica", "EPICA 3: Gestion de Casos Legales (exclusivo medico-abogado)"),
    ("HU-011","Listado de casos del medico","ver mis casos legales activos con su estado y prioridad","Medico","Must Have",3),
    ("HU-012","Creacion de un caso legal","registrar un nuevo caso legal con titulo, descripcion, prioridad y paciente","Medico","Must Have",3),
    ("HU-013","Detalle de un caso legal","ver el detalle completo de un caso con documentos y abogado asignado","Medico","Must Have",3),
    ("HU-014","Gestion de episodios clinicos (admin)","registrar y visualizar episodios clinicos asociados a pacientes (ingresos, altas)","Administrador","Should Have",4),
    ("HU-037","Asignacion de abogado a un caso legal","asignar un abogado directamente a mi caso para formalizar la relacion profesional","Medico","Should Have",5),
    ("HU-038","Edicion de notas de un caso legal","editar las notas de un caso para agregar contexto o actualizaciones","Medico","Could Have",5),
    # Epica 4
    ("epica", "EPICA 4: Vinculacion Medico-Abogado (Matching)"),
    ("HU-015","Recomendaciones de abogados para medico","ver una lista de abogados recomendados segun mi perfil con score de compatibilidad","Medico","Must Have",3),
    ("HU-016","Envio de solicitud de contacto","enviar una solicitud de contacto a un abogado con un mensaje personalizado","Medico","Must Have",3),
    ("HU-017","Gestion de solicitudes recibidas (abogado)","ver las solicitudes recibidas y poder aceptarlas o rechazarlas","Abogado","Must Have",3),
    ("HU-019","Directorio de abogados (vista medico)","explorar un directorio de abogados disponibles con sus perfiles profesionales","Medico","Should Have",3),
    ("HU-020","Directorio de medicos (vista abogado)","ver un directorio de los medicos registrados en la plataforma","Abogado","Should Have",3),
    ("HU-039","Mensaje de respuesta al gestionar solicitudes","incluir un mensaje al aceptar o rechazar una solicitud de contacto","Abogado","Could Have",6),
    # Epica 5
    ("epica", "EPICA 5: Gestion de Pacientes"),
    ("HU-021","Listado de pacientes","ver todos los pacientes registrados con busqueda por nombre o DNI","Administrador","Must Have",4),
    ("HU-022","Registro de nuevo paciente","registrar un nuevo paciente con sus datos personales y medicos","Administrador","Must Have",4),
    ("HU-023","Edicion de paciente","editar los datos de un paciente existente para mantener la informacion actualizada","Administrador","Should Have",4),
    # Epica 6
    ("epica", "EPICA 6: Gestion de Usuarios"),
    ("HU-024","Listado de usuarios del sistema","ver todos los usuarios del sistema con su rol y estado","Administrador","Must Have",4),
    ("HU-025","Creacion de nuevo usuario","registrar un nuevo usuario asignandole nombre, email y rol","Administrador","Must Have",4),
    ("HU-026","Activar/desactivar usuario","activar o desactivar un usuario sin eliminar su registro","Administrador","Should Have",4),
    # Epica 7
    ("epica", "EPICA 7: Auditoria y Trazabilidad"),
    ("HU-027","Registro de auditoria","ver un registro cronologico de todas las acciones del sistema","Administrador","Must Have",4),
    ("HU-043","Exportar listado de auditoria a CSV","exportar el registro de auditoria a un archivo CSV con los filtros aplicados","Administrador","Could Have",6),
    # Epica 8
    ("epica", "EPICA 8: Dashboards y Visualizacion"),
    ("HU-028","Dashboard del medico","ver un resumen de mis casos activos, documentos pendientes y recomendaciones","Medico","Must Have",2),
    ("HU-018","Dashboard del abogado","ver un resumen de solicitudes pendientes y casos activos","Abogado","Must Have",3),
    ("HU-029","Dashboard del administrador","ver estadisticas globales del sistema con usuarios, documentos y auditoria reciente","Administrador","Must Have",4),
    # Epica 9
    ("epica", "EPICA 9: Perfil de Usuario"),
    ("HU-030","Perfil del medico","ver y editar mi perfil profesional (CMP, especialidad, hospital, telefono, bio)","Medico","Should Have",5),
    ("HU-031","Perfil del abogado","ver y editar mi perfil profesional (CAB, especialidades, disponibilidad, bio)","Abogado","Should Have",5),
    ("HU-040","Toggle de disponibilidad del abogado","indicar si estoy disponible para aceptar nuevos casos desde mi perfil","Abogado","Could Have",6),
    # Epica 11
    ("epica", "EPICA 11: Firmas Digitales y Consentimiento Informado"),
    ("HU-034","Firma digital de documentos","registrar mi firma en un documento para validarlo de forma trazable","Medico","Should Have",5),
    ("HU-035","Gestion de consentimiento informado","generar y hacer seguimiento de registros de consentimiento asociados a un documento","Medico","Should Have",6),
    # Epica 12
    ("epica", "EPICA 12: Gestion de Archivos Adjuntos"),
    ("HU-036","Adjuntar documentos a un caso legal","adjuntar archivos (PDF, imagenes) a un caso para consolidar la evidencia","Medico","Should Have",6),
    # Epica 13
    ("epica", "EPICA 13: Evaluacion de Riesgo ML — Planificado (Sprint 7)"),
    ("HU-044","Evaluacion de riesgo medico-legal","recibir una estimacion de riesgo ML para mi caso con score y nivel (bajo/alto)","Medico","Must Have",7),
    ("HU-045","Visualizacion de factores XAI","ver una explicacion clara de los factores que generaron el riesgo (SHAP/LIME)","Medico","Must Have",7),
    ("HU-046","Historial de evaluaciones de riesgo","consultar el historial de evaluaciones de riesgo por paciente o caso","Medico","Should Have",7),
    ("HU-047","Recomendaciones automaticas por riesgo","recibir sugerencias de acciones mitigadoras basadas en el nivel de riesgo","Medico","Should Have",7),
]

prio_c2 = {"Must Have":(MUST,WHITE),"Should Have":(SHOULD,WHITE),"Could Have":(COULD,WHITE)}

R2 = HDR2 + 1
for entry in hu_data:
    if entry[0] == "epica":
        ws2.row_dimensions[R2].height = 18
        merge(ws2, R2, 1, R2, 6)
        cell(ws2, R2, 1, entry[1], bold=True, size=10, color=WHITE, bg=TEAL, align="left")
        R2 += 1
        continue

    hid, nombre, quiero, rol, prio, spr = entry
    ws2.row_dimensions[R2].height = 32
    bg = LTGRAY if R2 % 2 == 0 else WHITE

    cell(ws2,R2,1,hid,bold=True,size=9,color=TEAL,align="center",bg=bg).border=thin_border()
    cell(ws2,R2,2,nombre,bold=True,size=9,wrap=True,bg=bg).border=thin_border()
    cell(ws2,R2,3,quiero,size=9,wrap=True,color=DGRAY,bg=bg).border=thin_border()
    cell(ws2,R2,4,rol,size=9,align="center",color=DGRAY,bg=bg).border=thin_border()
    pbg2,pfg2=prio_c2.get(prio,(COULD,WHITE))
    cell(ws2,R2,5,prio,bold=True,size=8,color=pfg2,bg=pbg2,align="center").border=thin_border()
    cell(ws2,R2,6,f"S{spr}",size=9,color=DGRAY,align="center",bg=bg).border=thin_border()
    R2 += 1

ws2.freeze_panes = "A4"
ws2.auto_filter.ref = f"A{HDR2}:F{HDR2}"

wb.save(DST)
print("OK:", DST)
